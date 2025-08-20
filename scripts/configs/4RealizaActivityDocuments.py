import openpyxl
import mysql.connector
import uuid
import os
from datetime import datetime

# Configurações de conexão com o banco de dados MySQL
DB_CONFIG = {
    "host": "35.184.183.88",
    "port": 3306,
    "user": "veridion-admin-user",
    "password": "uMsgC-x+uAA]yRG1",
    "database": "realiza_mysql_development"
}
# DB_CONFIG = {
#     "host": "177.170.30.9",
#     "port": 8004,
#     "user": "veridion_user",
#     "password": "SenhaSegura123!",
#     "database": "dbrealiza"
# }


def importar_atividades_e_documentos(path_excel):
    wb = openpyxl.load_workbook(path_excel, data_only=True)
    ws = wb["Planilha1"]  # Nome da aba que contém os dados

    atividades_documentos = []
    atividade_atual = None

    # Percorrendo as linhas da planilha
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):  # Colunas A e B
        atividade = row[0].value
        documento = row[1].value

        # Se a célula da coluna A não está vazia, é uma nova atividade
        if atividade:
            atividade_atual = atividade
        # Caso contrário, o documento é associado à atividade anterior
        if documento and atividade_atual:
            atividades_documentos.append({
                "atividade": atividade_atual,
                "documento": documento
            })

    return atividades_documentos


def processar_atividades_e_documentos(atividades_documentos):
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor()
        print("Conectado ao banco de dados com sucesso.")

        for item in atividades_documentos:
            atividade = item["atividade"]
            documento = item["documento"]

            # Inserir na tabela activity_repo (se a atividade não existir)
            cursor.execute(
                "SELECT id_activity FROM activity_repo WHERE title = %s", (atividade,)
            )
            atividade_row = cursor.fetchone()

            if not atividade_row:
                atividade_id = str(uuid.uuid4())  # Gerar um UUID único para a atividade
                cursor.execute(
                    "INSERT INTO activity_repo (id_activity, creation_date, risk, title) VALUES (%s, %s, %s, %s)",
                    (atividade_id, datetime.now(), 0, atividade)  # '0' para risco padrão
                )
                print(f"Atividade '{atividade}' inserida na tabela activity_repo.")
            else:
                atividade_id = atividade_row[0]  # Caso a atividade já exista, usa o ID existente

            # Procurar o id do documento na tabela document_matrix
            cursor.execute("SELECT id_document FROM document_matrix WHERE name = %s", (documento,))
            documento_row = cursor.fetchone()

            if documento_row:
                documento_id = documento_row[0]
                # Inserir na tabela activity_documents_repo
                try:
                    cursor.execute(
                        "INSERT INTO activity_documents_repo (id, id_activity, id_document) VALUES (%s, %s, %s)",
                        (str(uuid.uuid4()), atividade_id, documento_id)
                    )
                    print(f"Documento '{documento}' associado à atividade '{atividade}'.")
                except mysql.connector.Error as e:
                    # Ignorar erro de duplicação
                    if e.errno == 1062:  # Erro de duplicação
                        print(f"Documento '{documento}' já associado à atividade '{atividade}', ignorado.")
                    else:
                        raise e  # Levanta outros erros
            else:
                print(f"Documento '{documento}' não encontrado na tabela document_matrix.")

        conn.commit()
        print("Todas as atividades e documentos foram inseridos com sucesso.")

    except mysql.connector.Error as err:
        print(f"Erro no banco de dados: {err}")
    finally:
        if conn:
            cursor.close()
            conn.close()
            print("Conexão com o banco encerrada.")


def main():
    base_dir = os.path.dirname(__file__)
    path = os.path.join(base_dir, "BL_DOCTOS POR ATIVIDADE_IGUA.xlsx")
    atividades_documentos = importar_atividades_e_documentos(path)
    processar_atividades_e_documentos(atividades_documentos)


if __name__ == "__main__":
    main()
